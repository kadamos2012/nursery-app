import os
from datetime import datetime, date, time as dtime

from flask import Flask, request, jsonify, render_template, redirect, url_for, session
from flask_login import (
    LoginManager, login_user, logout_user, login_required, current_user
)
from flask_cors import CORS

from models import (
    db, Nursery, SchoolClass, Teacher, Parent, Child, ParentChild,
    DailyLog, AttendanceRecord, Payment, Message, PushSubscription,
    Owner, Employee, SalaryPayment, Expense, Activity, Addon, ChildAddon,
    PaymentMethod, Advance, ExpenseCategory, Announcement,
    Trip, TripCostItem, TripClass, TripRegistration,
    EnrollmentRequest, ClassPhoto
)
from pywebpush import webpush, WebPushException
import json as json_lib
import base64
import secrets
import requests
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-change-me")

db_url = os.environ.get("DATABASE_URL", "sqlite:///nursery.db")
# Render's Postgres URLs start with "postgres://"; SQLAlchemy needs "postgresql://"
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

VAPID_PUBLIC_KEY = os.environ.get("VAPID_PUBLIC_KEY", "")
VAPID_PRIVATE_KEY = os.environ.get("VAPID_PRIVATE_KEY", "")
VAPID_CLAIMS = {"sub": os.environ.get("VAPID_CONTACT_EMAIL", "mailto:admin@example.com")}

# Needed so the browser-based frontend (a different origin) can send the
# session cookie back on each request once this is deployed.
app.config["SESSION_COOKIE_SAMESITE"] = "None"
app.config["SESSION_COOKIE_SECURE"] = True

db.init_app(app)

# ---------------------------------------------------------------------------
# Telegram bot (free alternative notification channel)
# ---------------------------------------------------------------------------
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_BOT_USERNAME = ""  # fetched automatically at startup via getMe
BACKEND_URL = os.environ.get("BACKEND_URL") or os.environ.get("RENDER_EXTERNAL_URL", "")


def send_telegram_message(chat_id, text):
    if not TELEGRAM_BOT_TOKEN:
        return
    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
            json={"chat_id": chat_id, "text": text},
            timeout=8,
        )
    except Exception:
        pass  # notifications are best-effort


def setup_telegram_webhook():
    """Registers our webhook URL with Telegram and fetches the bot's username.
    Safe to call on every startup — Telegram just re-confirms the same URL."""
    global TELEGRAM_BOT_USERNAME
    if not TELEGRAM_BOT_TOKEN:
        return
    try:
        resp = requests.get(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getMe", timeout=8)
        TELEGRAM_BOT_USERNAME = resp.json().get("result", {}).get("username", "")
    except Exception:
        pass

    if BACKEND_URL:
        try:
            requests.post(
                f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/setWebhook",
                json={"url": f"{BACKEND_URL.rstrip('/')}/telegram/webhook"},
                timeout=8,
            )
        except Exception:
            pass


with app.app_context():
    setup_telegram_webhook()

# Allow the mobile web app (served from a different domain, e.g. claude.ai
# artifacts or your own frontend host) to call this API with cookies.
CORS(app, supports_credentials=True, origins=os.environ.get("ALLOWED_ORIGIN", "*"))

login_manager = LoginManager()
login_manager.init_app(app)


@login_manager.user_loader
def load_user(user_id):
    kind, raw_id = user_id.split(":")
    if kind == "teacher":
        return Teacher.query.get(int(raw_id))
    if kind == "parent":
        return Parent.query.get(int(raw_id))
    if kind == "owner":
        return Owner.query.get(int(raw_id))
    return None


def current_role():
    if not current_user.is_authenticated:
        return None
    if isinstance(current_user, Teacher):
        return "teacher"
    if isinstance(current_user, Owner):
        return "owner"
    return "parent"


def compute_child_expected_fee(child):
    addons_total = sum(float(ca.addon.monthly_fee) for ca in ChildAddon.query.filter_by(child_id=child.id).all())
    subtotal = float(child.monthly_fee) + addons_total
    if child.discount_type == "fixed":
        discount = float(child.discount_value)
    elif child.discount_type == "percent":
        discount = subtotal * float(child.discount_value) / 100
    else:
        discount = 0
    return max(subtotal - discount, 0)


def child_belongs_to_parent(child_id, parent_id):
    return ParentChild.query.filter_by(child_id=child_id, parent_id=parent_id).first() is not None


def notify_child_parents(child_id, title, body):
    """Sends a Web Push notification and/or a Telegram message to every parent
    linked to this child, depending on what each parent has set up."""
    links = ParentChild.query.filter_by(child_id=child_id).all()
    parent_ids = [l.parent_id for l in links]

    if VAPID_PRIVATE_KEY:
        subs = PushSubscription.query.filter(PushSubscription.parent_id.in_(parent_ids)).all()
        for sub in subs:
            try:
                webpush(
                    subscription_info={
                        "endpoint": sub.endpoint,
                        "keys": {"p256dh": sub.p256dh, "auth": sub.auth},
                    },
                    data=json_lib.dumps({"title": title, "body": body}),
                    vapid_private_key=VAPID_PRIVATE_KEY,
                    vapid_claims=dict(VAPID_CLAIMS),
                )
            except WebPushException:
                db.session.delete(sub)
        db.session.commit()

    telegram_parents = Parent.query.filter(
        Parent.id.in_(parent_ids), Parent.telegram_chat_id.isnot(None)
    ).all()
    for p in telegram_parents:
        send_telegram_message(p.telegram_chat_id, f"{title}\n{body}")


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

@app.route("/telegram/webhook", methods=["POST"])
def telegram_webhook():
    update = request.get_json(silent=True) or {}
    message = update.get("message", {})
    text = message.get("text", "")
    chat_id = message.get("chat", {}).get("id")

    if text.startswith("/start") and chat_id:
        parts = text.split(maxsplit=1)
        if len(parts) == 2:
            code = parts[1].strip()
            parent = Parent.query.filter_by(telegram_link_code=code).first()
            if parent:
                parent.telegram_chat_id = str(chat_id)
                db.session.commit()
                send_telegram_message(chat_id, f"تم الربط بنجاح ✅ هتوصلك هنا كل تحديثات {parent.name.replace('ولية أمر ', '').replace('والد ', '')} من الحضانة.")
            else:
                send_telegram_message(chat_id, "الكود مش صحيح أو منتهي. جربي تربطي تاني من التطبيق.")
        else:
            send_telegram_message(chat_id, "أهلاً 👋 افتحي تطبيق الحضانة واضغطي على 'ربط تليجرام' عشان نوصلك بحسابك.")

    return {"ok": True}


@app.route("/api/parent/telegram-link")
@login_required
def api_parent_telegram_link():
    if current_role() != "parent":
        return jsonify({"error": "غير مصرح"}), 403

    if current_user.telegram_chat_id:
        return jsonify({"linked": True})

    if not current_user.telegram_link_code:
        current_user.telegram_link_code = secrets.token_urlsafe(8)
        db.session.commit()

    bot_username = TELEGRAM_BOT_USERNAME
    return jsonify({
        "linked": False,
        "link": f"https://t.me/{bot_username}?start={current_user.telegram_link_code}" if bot_username else None,
    })


@app.route("/api/auth/login", methods=["POST"])
def api_login():
    """
    Body: { "role": "parent" | "teacher", "phone": "...", "password": "..." }
    """
    data = request.get_json(force=True)
    role = data.get("role")
    phone = data.get("phone", "").strip()
    password = data.get("password", "")

    model = Teacher if role == "teacher" else Parent
    user = model.query.filter_by(phone=phone).first()

    if not user or not user.check_password(password):
        return jsonify({"error": "رقم الهاتف أو كلمة المرور غير صحيحة"}), 401

    login_user(user)
    return jsonify({"ok": True, "role": role, "name": user.name, "id": user.id})


@app.route("/api/auth/logout", methods=["POST"])
@login_required
def api_logout():
    logout_user()
    return jsonify({"ok": True})


@app.route("/api/auth/me")
def api_me():
    if not current_user.is_authenticated:
        return jsonify({"authenticated": False})
    return jsonify({
        "authenticated": True,
        "role": current_role(),
        "id": current_user.id,
        "name": current_user.name,
    })


# ---------------------------------------------------------------------------
# Parent-facing API (consumed by the mobile web app)
# ---------------------------------------------------------------------------

@app.route("/api/parent/children")
@login_required
def parent_children():
    if current_role() != "parent":
        return jsonify({"error": "غير مصرح"}), 403

    links = ParentChild.query.filter_by(parent_id=current_user.id).all()
    children = []
    for link in links:
        c = Child.query.get(link.child_id)
        children.append({
            "id": c.id,
            "name": c.name,
            "class_name": c.school_class.name,
            "photo_url": url_for("child_photo", child_id=c.id, _external=True) if c.photo_data else None,
            "medical_notes": c.medical_notes,
        })
    return jsonify(children)


@app.route("/api/child/<int:child_id>/today")
@login_required
def child_today(child_id):
    if current_role() == "parent" and not child_belongs_to_parent(child_id, current_user.id):
        return jsonify({"error": "غير مصرح"}), 403

    log = DailyLog.query.filter_by(child_id=child_id, date=date.today()).first()
    if not log:
        return jsonify({"exists": False})

    return jsonify({
        "exists": True,
        "date": log.date.isoformat(),
        "arrival_time": log.arrival_time.isoformat() if log.arrival_time else None,
        "departure_time": log.departure_time.isoformat() if log.departure_time else None,
        "meal_status": log.meal_status,
        "nap_minutes": log.nap_minutes,
        "mood": log.mood,
        "note": log.note,
        "photo_url": log.photo_url,
    })


@app.route("/api/child/<int:child_id>/logs")
@login_required
def child_logs(child_id):
    """History of daily logs, most recent first. ?limit=30"""
    if current_role() == "parent" and not child_belongs_to_parent(child_id, current_user.id):
        return jsonify({"error": "غير مصرح"}), 403

    limit = int(request.args.get("limit", 30))
    logs = (DailyLog.query.filter_by(child_id=child_id)
            .order_by(DailyLog.date.desc()).limit(limit).all())

    return jsonify([{
        "date": l.date.isoformat(),
        "meal_status": l.meal_status,
        "nap_minutes": l.nap_minutes,
        "mood": l.mood,
        "note": l.note,
        "photo_url": l.photo_url,
    } for l in logs])


@app.route("/api/child/<int:child_id>/attendance")
@login_required
def child_attendance(child_id):
    """?year=2026&month=8"""
    if current_role() == "parent" and not child_belongs_to_parent(child_id, current_user.id):
        return jsonify({"error": "غير مصرح"}), 403

    year = int(request.args.get("year", date.today().year))
    month = int(request.args.get("month", date.today().month))

    records = AttendanceRecord.query.filter(
        AttendanceRecord.child_id == child_id,
        db.extract("year", AttendanceRecord.date) == year,
        db.extract("month", AttendanceRecord.date) == month,
    ).all()

    return jsonify([{"date": r.date.isoformat(), "present": r.present} for r in records])


@app.route("/api/child/<int:child_id>/payments")
@login_required
def child_payments(child_id):
    if current_role() == "parent" and not child_belongs_to_parent(child_id, current_user.id):
        return jsonify({"error": "غير مصرح"}), 403

    payments = Payment.query.filter_by(child_id=child_id).order_by(
        Payment.year.desc(), Payment.month.desc()
    ).all()

    return jsonify([{
        "month": p.month, "year": p.year, "amount": str(p.amount),
        "paid": p.paid, "paid_date": p.paid_date.isoformat() if p.paid_date else None,
    } for p in payments])


@app.route("/api/child/<int:child_id>/payment-status")
@login_required
def child_payment_status(child_id):
    if current_role() == "parent" and not child_belongs_to_parent(child_id, current_user.id):
        return jsonify({"error": "غير مصرح"}), 403

    child = Child.query.get_or_404(child_id)
    today = date.today()
    payment = Payment.query.filter_by(child_id=child_id, month=today.month, year=today.year).first()

    if payment:
        return jsonify({
            "month": today.month, "year": today.year,
            "amount": str(payment.amount), "paid": payment.paid,
        })

    expected = compute_child_expected_fee(child)
    return jsonify({"month": today.month, "year": today.year, "amount": str(expected), "paid": False})


@app.route("/api/child/<int:child_id>/announcements")
@login_required
def child_announcements(child_id):
    if current_role() == "parent" and not child_belongs_to_parent(child_id, current_user.id):
        return jsonify({"error": "غير مصرح"}), 403

    child = Child.query.get_or_404(child_id)
    announcements = Announcement.query.filter(
        (Announcement.class_id.is_(None)) | (Announcement.class_id == child.class_id)
    ).order_by(Announcement.created_at.desc()).limit(10).all()

    return jsonify([{
        "id": a.id, "title": a.title, "body": a.body,
        "created_at": a.created_at.isoformat(),
        "scope": "كل الحضانة" if not a.class_id else a.school_class.name,
    } for a in announcements])


@app.route("/api/child/<int:child_id>/trips")
@login_required
def child_trips(child_id):
    if current_role() == "parent" and not child_belongs_to_parent(child_id, current_user.id):
        return jsonify({"error": "غير مصرح"}), 403

    child = Child.query.get_or_404(child_id)
    eligible_trip_ids = [tc.trip_id for tc in TripClass.query.filter_by(class_id=child.class_id).all()]
    trips = Trip.query.filter(
        Trip.id.in_(eligible_trip_ids), Trip.status == "published"
    ).order_by(Trip.date).all() if eligible_trip_ids else []

    result = []
    for t in trips:
        reg = TripRegistration.query.filter_by(trip_id=t.id, child_id=child_id).first()
        result.append({
            "id": t.id, "title": t.title, "description": t.description,
            "date": t.date.isoformat() if t.date else None,
            "price": str(t.final_price or 0),
            "registration": {
                "paid": reg.paid, "requested_by": reg.requested_by,
            } if reg else None,
        })
    return jsonify(result)


@app.route("/api/child/<int:child_id>/trips/<int:trip_id>/register", methods=["POST"])
@login_required
def child_register_trip(child_id, trip_id):
    if current_role() == "parent" and not child_belongs_to_parent(child_id, current_user.id):
        return jsonify({"error": "غير مصرح"}), 403

    trip = Trip.query.get_or_404(trip_id)
    existing = TripRegistration.query.filter_by(trip_id=trip_id, child_id=child_id).first()
    if existing:
        return jsonify({"ok": True, "already_registered": True})

    db.session.add(TripRegistration(
        trip_id=trip_id, child_id=child_id, price=trip.final_price or 0, requested_by="parent",
    ))
    db.session.commit()
    return jsonify({"ok": True})


@app.route("/api/child/<int:child_id>/class-photos")
@login_required
def child_class_photos(child_id):
    if current_role() == "parent" and not child_belongs_to_parent(child_id, current_user.id):
        return jsonify({"error": "غير مصرح"}), 403

    child = Child.query.get_or_404(child_id)
    photos = ClassPhoto.query.filter_by(class_id=child.class_id).order_by(ClassPhoto.created_at.desc()).limit(20).all()
    return jsonify([{
        "id": p.id, "url": url_for("serve_class_photo", photo_id=p.id, _external=True),
        "caption": p.caption, "created_at": p.created_at.isoformat(),
    } for p in photos])


@app.route("/api/child/<int:child_id>/messages", methods=["GET", "POST"])
@login_required
def child_messages(child_id):
    if current_role() == "parent" and not child_belongs_to_parent(child_id, current_user.id):
        return jsonify({"error": "غير مصرح"}), 403

    if request.method == "POST":
        data = request.get_json(force=True)
        msg = Message(
            child_id=child_id,
            sender_type=current_role(),
            sender_id=current_user.id,
            text=data.get("text", "").strip(),
        )
        db.session.add(msg)
        db.session.commit()

        if current_role() == "teacher":
            child = Child.query.get(child_id)
            notify_child_parents(child_id, f"رسالة جديدة عن {child.name}", msg.text[:100])

        return jsonify({"ok": True, "id": msg.id})

    msgs = Message.query.filter_by(child_id=child_id).order_by(Message.timestamp).all()
    return jsonify([{
        "sender_type": m.sender_type,
        "text": m.text,
        "timestamp": m.timestamp.isoformat(),
    } for m in msgs])


# ---------------------------------------------------------------------------
# Push notifications (Web Push API)
# ---------------------------------------------------------------------------

@app.route("/api/push/vapid-public-key")
def push_public_key():
    return jsonify({"publicKey": VAPID_PUBLIC_KEY})


@app.route("/api/push/subscribe", methods=["POST"])
@login_required
def push_subscribe():
    if current_role() != "parent":
        return jsonify({"error": "غير مصرح"}), 403

    data = request.get_json(force=True)
    endpoint = data.get("endpoint")
    keys = data.get("keys", {})

    existing = PushSubscription.query.filter_by(endpoint=endpoint).first()
    if existing:
        existing.p256dh = keys.get("p256dh")
        existing.auth = keys.get("auth")
    else:
        db.session.add(PushSubscription(
            parent_id=current_user.id, endpoint=endpoint,
            p256dh=keys.get("p256dh"), auth=keys.get("auth"),
        ))
    db.session.commit()
    return jsonify({"ok": True})


@app.route("/api/push/unsubscribe", methods=["POST"])
@login_required
def push_unsubscribe():
    data = request.get_json(force=True)
    PushSubscription.query.filter_by(endpoint=data.get("endpoint")).delete()
    db.session.commit()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Teacher-facing quick-entry pages (server-rendered, built for speed on a phone)
# ---------------------------------------------------------------------------

FRONTEND_URL = os.environ.get("FRONTEND_URL", "https://nursery-app-frontend-i3jq.onrender.com")


@app.route("/")
def index():
    return redirect(url_for("unified_login"))


@app.route("/login", methods=["GET", "POST"])
def unified_login():
    if request.method == "POST":
        phone = request.form.get("phone", "").strip()
        password = request.form.get("password", "")

        owner = Owner.query.filter_by(phone=phone).first()
        if owner and owner.check_password(password):
            login_user(owner)
            return redirect(url_for("owner_dashboard"))

        teacher = Teacher.query.filter_by(phone=phone).first()
        if teacher and teacher.check_password(password):
            login_user(teacher)
            return redirect(url_for("teacher_dashboard"))

        parent = Parent.query.filter_by(phone=phone).first()
        if parent and parent.check_password(password):
            login_user(parent)
            return redirect(FRONTEND_URL)

        return render_template("login.html", error="رقم الهاتف أو كلمة المرور غير صحيحة", nursery=Nursery.query.first())

    return render_template("login.html", error=None, nursery=Nursery.query.first())


@app.route("/teacher/photos", methods=["GET", "POST"])
@login_required
def teacher_class_photos():
    if current_role() != "teacher":
        return redirect(url_for("unified_login"))

    class_id = current_user.class_id
    if request.method == "POST":
        photo_file = request.files.get("photo")
        caption = request.form.get("caption", "").strip()
        if photo_file and photo_file.filename and class_id:
            raw = photo_file.read()
            if len(raw) <= 3 * 1024 * 1024:
                db.session.add(ClassPhoto(
                    class_id=class_id, photo_data=base64.b64encode(raw).decode("ascii"),
                    photo_mime=photo_file.mimetype or "image/jpeg", caption=caption,
                    uploaded_by_teacher_id=current_user.id,
                ))
                db.session.commit()
        return redirect(url_for("teacher_class_photos"))

    photos = ClassPhoto.query.filter_by(class_id=class_id).order_by(ClassPhoto.created_at.desc()).all() if class_id else []
    return render_template("teacher_photos.html", photos=photos)


@app.route("/class-photo/<int:photo_id>")
@login_required
def serve_class_photo(photo_id):
    photo = ClassPhoto.query.get_or_404(photo_id)
    if current_role() == "parent":
        # must have a child in that class
        child_ids = [pc.child_id for pc in ParentChild.query.filter_by(parent_id=current_user.id).all()]
        classes_ok = Child.query.filter(Child.id.in_(child_ids), Child.class_id == photo.class_id).first()
        if not classes_ok:
            return "غير مصرح", 403
    raw = base64.b64decode(photo.photo_data)
    return app.response_class(raw, mimetype=photo.photo_mime)


@app.route("/teacher")
@login_required
def teacher_dashboard():
    if current_role() != "teacher":
        return redirect(url_for("unified_login"))
    if current_user.class_id:
        children = Child.query.filter_by(class_id=current_user.class_id, archived=False).all()
    else:
        children = Child.query.filter_by(archived=False).all()

    today = date.today()
    logs_today = {l.child_id: l for l in DailyLog.query.filter_by(date=today).all()}

    return render_template("teacher_dashboard.html", children=children, logs_today=logs_today, today=today)


@app.route("/teacher/child/<int:child_id>/log", methods=["GET", "POST"])
@login_required
def teacher_child_log(child_id):
    if current_role() != "teacher":
        return redirect(url_for("unified_login"))

    child = Child.query.get_or_404(child_id)
    today = date.today()
    log = DailyLog.query.filter_by(child_id=child_id, date=today).first()
    if not log:
        log = DailyLog(child_id=child_id, date=today, created_by_teacher_id=current_user.id)
        db.session.add(log)

    if request.method == "POST":
        log.meal_status = request.form.get("meal_status") or log.meal_status
        log.mood = request.form.get("mood") or log.mood
        note = request.form.get("note")
        if note:
            log.note = note
        nap = request.form.get("nap_minutes")
        if nap:
            log.nap_minutes = int(nap)
        log.updated_at = datetime.utcnow()
        db.session.commit()
        notify_child_parents(child_id, f"تحديث جديد لـ {child.name}", "المعلمة حدّثت يوميات طفلك — افتحي التطبيق للتفاصيل")
        return redirect(url_for("teacher_dashboard"))

    return render_template("teacher_child_log.html", child=child, log=log)


# ---------------------------------------------------------------------------
# Owner-facing pages: accounting, classes, students, activities, staff
# ---------------------------------------------------------------------------

def require_owner():
    return current_role() == "owner"


@app.route("/teacher/birthdays")
@login_required
def teacher_birthdays():
    if current_role() != "teacher":
        return redirect(url_for("unified_login"))

    today = date.today()
    if current_user.class_id:
        children = Child.query.filter_by(class_id=current_user.class_id, archived=False).all()
    else:
        children = Child.query.filter_by(archived=False).all()

    birthday_kids = [c for c in children if c.birth_date and c.birth_date.month == today.month]
    birthday_kids.sort(key=lambda c: c.birth_date.day)

    return render_template("teacher_birthdays.html", children=birthday_kids, today=today)


@app.route("/apply", methods=["GET", "POST"])
def public_apply():
    if request.method == "POST":
        nursery = Nursery.query.first()
        parent_name = request.form.get("parent_name", "").strip()
        phone = request.form.get("phone", "").strip()
        child_name = request.form.get("child_name", "").strip()
        child_birth_date = request.form.get("child_birth_date") or None
        notes = request.form.get("notes", "").strip()

        if parent_name and phone:
            db.session.add(EnrollmentRequest(
                nursery_id=nursery.id if nursery else 1, parent_name=parent_name, phone=phone,
                child_name=child_name, child_birth_date=child_birth_date, notes=notes,
            ))
            db.session.commit()
            return render_template("public_apply.html", nursery=nursery, submitted=True)

    nursery = Nursery.query.first()
    return render_template("public_apply.html", nursery=nursery, submitted=False)


@app.route("/owner/enrollment-requests", methods=["GET", "POST"])
@login_required
def owner_enrollment_requests():
    if not require_owner():
        return redirect(url_for("unified_login"))

    if request.method == "POST":
        request_id = request.form.get("request_id")
        new_status = request.form.get("status")
        req = EnrollmentRequest.query.get_or_404(request_id)
        req.status = new_status
        db.session.commit()
        return redirect(url_for("owner_enrollment_requests"))

    requests_list = EnrollmentRequest.query.order_by(EnrollmentRequest.created_at.desc()).all()
    return render_template("owner_enrollment_requests.html", requests_list=requests_list)


@app.route("/owner/settings", methods=["GET", "POST"])
@login_required
def owner_settings():
    if not require_owner():
        return redirect(url_for("unified_login"))

    nursery = Nursery.query.get_or_404(current_user.nursery_id)

    if request.method == "POST":
        nursery.facebook_url = request.form.get("facebook_url", "").strip() or None
        nursery.instagram_url = request.form.get("instagram_url", "").strip() or None
        nursery.tiktok_url = request.form.get("tiktok_url", "").strip() or None

        logo_file = request.files.get("logo")
        if logo_file and logo_file.filename:
            raw = logo_file.read()
            if len(raw) > 2 * 1024 * 1024:
                return render_template("owner_settings.html", nursery=nursery, error="حجم اللوجو كبير أوي، اختاري صورة أصغر من 2 ميجا")
            nursery.logo_data = base64.b64encode(raw).decode("ascii")
            nursery.logo_mime = logo_file.mimetype or "image/png"

        db.session.commit()
        return redirect(url_for("owner_settings"))

    return render_template("owner_settings.html", nursery=nursery, error=None)


@app.route("/nursery/logo")
def nursery_logo():
    nursery = Nursery.query.first()
    if not nursery or not nursery.logo_data:
        return "", 404
    raw = base64.b64decode(nursery.logo_data)
    return app.response_class(raw, mimetype=nursery.logo_mime or "image/png")


@app.route("/api/nursery/branding")
def api_nursery_branding():
    nursery = Nursery.query.first()
    if not nursery:
        return jsonify({})
    return jsonify({
        "name": nursery.name,
        "logo_url": url_for("nursery_logo", _external=True) if nursery.logo_data else None,
        "facebook_url": nursery.facebook_url,
        "instagram_url": nursery.instagram_url,
        "tiktok_url": nursery.tiktok_url,
    })


@app.route("/owner")
@login_required
def owner_dashboard():
    if not require_owner():
        return redirect(url_for("unified_login"))

    today = date.today()
    month, year = today.month, today.year

    tuition_paid = db.session.query(db.func.coalesce(db.func.sum(Payment.amount), 0)).filter(
        Payment.month == month, Payment.year == year, Payment.paid.is_(True)
    ).scalar()
    tuition_due = db.session.query(db.func.coalesce(db.func.sum(Payment.amount), 0)).filter(
        Payment.month == month, Payment.year == year, Payment.paid.is_(False)
    ).scalar()
    expenses_total = db.session.query(db.func.coalesce(db.func.sum(Expense.amount), 0)).filter(
        db.extract("month", Expense.date) == month, db.extract("year", Expense.date) == year
    ).scalar()
    salaries_total = db.session.query(db.func.coalesce(db.func.sum(SalaryPayment.amount), 0)).filter(
        SalaryPayment.month == month, SalaryPayment.year == year
    ).scalar()
    activities_cost = db.session.query(db.func.coalesce(db.func.sum(Activity.cost), 0)).filter(
        db.extract("month", Activity.date) == month, db.extract("year", Activity.date) == year
    ).scalar()

    net = float(tuition_paid) - float(expenses_total) - float(salaries_total) - float(activities_cost)

    classes_count = SchoolClass.query.count()
    students_count = Child.query.count()
    staff_count = Employee.query.filter_by(active=True).count()

    return render_template(
        "owner_dashboard.html", month=month, year=year,
        tuition_paid=tuition_paid, tuition_due=tuition_due, expenses_total=expenses_total,
        salaries_total=salaries_total, activities_cost=activities_cost, net=net,
        classes_count=classes_count, students_count=students_count, staff_count=staff_count,
    )


@app.route("/owner/announcements", methods=["GET", "POST"])
@login_required
def owner_announcements():
    if not require_owner():
        return redirect(url_for("unified_login"))

    if request.method == "POST":
        title = request.form.get("title", "").strip()
        body = request.form.get("body", "").strip()
        class_id = request.form.get("class_id") or None
        if title and body:
            db.session.add(Announcement(
                nursery_id=current_user.nursery_id, class_id=class_id,
                title=title, body=body, created_by_owner_id=current_user.id,
            ))
            db.session.commit()
        return redirect(url_for("owner_announcements"))

    classes = SchoolClass.query.all()
    announcements = Announcement.query.order_by(Announcement.created_at.desc()).all()
    return render_template("owner_announcements.html", classes=classes, announcements=announcements)


@app.route("/owner/announcements/<int:announcement_id>/delete", methods=["POST"])
@login_required
def owner_delete_announcement(announcement_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    announcement = Announcement.query.get_or_404(announcement_id)
    db.session.delete(announcement)
    db.session.commit()
    return redirect(url_for("owner_announcements"))


def compute_trip_pricing(trip):
    """Returns (cost_per_child, suggested_price) based on cost items, estimated
    students, and profit margin."""
    cost_per_child = 0.0
    for item in trip.cost_items:
        if item.cost_type == "per_child":
            cost_per_child += float(item.amount)
        else:
            cost_per_child += float(item.amount) / max(trip.estimated_students, 1)
    suggested_price = cost_per_child * (1 + float(trip.profit_margin_percent) / 100)
    return cost_per_child, suggested_price


@app.route("/owner/trips")
@login_required
def owner_trips():
    if not require_owner():
        return redirect(url_for("unified_login"))

    trips = Trip.query.order_by(Trip.created_at.desc()).all()
    trip_data = []
    for t in trips:
        cost_per_child, suggested = compute_trip_pricing(t)
        registered = len(t.registrations)
        paid_count = len([r for r in t.registrations if r.paid])
        trip_data.append({
            "trip": t, "cost_per_child": cost_per_child, "suggested": suggested,
            "registered": registered, "paid_count": paid_count,
        })
    return render_template("owner_trips.html", trip_data=trip_data)


@app.route("/owner/trips/new", methods=["GET", "POST"])
@login_required
def owner_new_trip():
    if not require_owner():
        return redirect(url_for("unified_login"))

    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        trip_date = request.form.get("date") or None
        estimated_students = int(request.form.get("estimated_students") or 1)
        margin = request.form.get("profit_margin_percent") or 0

        trip = Trip(
            nursery_id=current_user.nursery_id, title=title, description=description,
            date=trip_date, estimated_students=estimated_students, profit_margin_percent=margin,
        )
        db.session.add(trip)
        db.session.flush()

        names = request.form.getlist("item_name")
        amounts = request.form.getlist("item_amount")
        types = request.form.getlist("item_type")
        for name, amount, ctype in zip(names, amounts, types):
            if name.strip() and amount:
                db.session.add(TripCostItem(trip_id=trip.id, name=name.strip(), amount=amount, cost_type=ctype))

        db.session.commit()
        return redirect(url_for("owner_trip_detail", trip_id=trip.id))

    return render_template("owner_trip_new.html")


@app.route("/owner/trips/<int:trip_id>")
@login_required
def owner_trip_detail(trip_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    trip = Trip.query.get_or_404(trip_id)
    cost_per_child, suggested = compute_trip_pricing(trip)
    classes = SchoolClass.query.all()
    eligible_ids = {tc.class_id for tc in trip.eligible_classes}

    registrations = TripRegistration.query.filter_by(trip_id=trip_id).all()
    methods = PaymentMethod.query.filter_by(active=True).all()

    eligible_children = []
    if eligible_ids:
        eligible_children = Child.query.filter(Child.class_id.in_(eligible_ids), Child.archived.is_(False)).all()
    registered_child_ids = {r.child_id for r in registrations}
    unregistered_children = [c for c in eligible_children if c.id not in registered_child_ids]

    revenue_collected = sum(float(r.price) for r in registrations if r.paid)
    real_profit = (revenue_collected - float(trip.actual_total_cost)) if trip.actual_total_cost is not None else None

    return render_template(
        "owner_trip_detail.html", trip=trip, cost_per_child=cost_per_child, suggested=suggested,
        classes=classes, eligible_ids=eligible_ids, registrations=registrations, methods=methods,
        unregistered_children=unregistered_children, revenue_collected=revenue_collected, real_profit=real_profit,
    )


@app.route("/owner/trips/<int:trip_id>/publish", methods=["POST"])
@login_required
def owner_publish_trip(trip_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    trip = Trip.query.get_or_404(trip_id)
    trip.final_price = request.form.get("final_price") or 0
    trip_date = request.form.get("date")
    if trip_date:
        trip.date = trip_date

    TripClass.query.filter_by(trip_id=trip_id).delete()
    class_ids = request.form.getlist("class_ids")
    for cid in class_ids:
        db.session.add(TripClass(trip_id=trip_id, class_id=int(cid)))

    trip.status = "published"
    db.session.commit()
    return redirect(url_for("owner_trip_detail", trip_id=trip_id))


@app.route("/owner/trips/<int:trip_id>/register", methods=["POST"])
@login_required
def owner_register_child_trip(trip_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    trip = Trip.query.get_or_404(trip_id)
    child_id = request.form.get("child_id")
    if child_id:
        existing = TripRegistration.query.filter_by(trip_id=trip_id, child_id=child_id).first()
        if not existing:
            db.session.add(TripRegistration(
                trip_id=trip_id, child_id=child_id, price=trip.final_price or 0, requested_by="owner",
            ))
            db.session.commit()
    return redirect(url_for("owner_trip_detail", trip_id=trip_id))


@app.route("/owner/trips/<int:trip_id>/close", methods=["POST"])
@login_required
def owner_close_trip(trip_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    trip = Trip.query.get_or_404(trip_id)
    trip.actual_total_cost = request.form.get("actual_total_cost") or 0
    db.session.commit()
    return redirect(url_for("owner_trip_detail", trip_id=trip_id))


@app.route("/owner/trips/registrations/<int:reg_id>/pay", methods=["POST"])
@login_required
def owner_pay_trip_registration(reg_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    reg = TripRegistration.query.get_or_404(reg_id)
    reg.paid = True
    reg.paid_date = date.today()
    reg.payment_method_id = request.form.get("payment_method_id") or None
    db.session.commit()
    return redirect(url_for("owner_trip_detail", trip_id=reg.trip_id))


@app.route("/owner/birthdays")
@login_required
def owner_birthdays():
    if not require_owner():
        return redirect(url_for("unified_login"))

    today = date.today()
    children = Child.query.filter_by(archived=False).all()
    birthday_kids = [c for c in children if c.birth_date and c.birth_date.month == today.month]
    birthday_kids.sort(key=lambda c: c.birth_date.day)

    return render_template("owner_birthdays.html", children=birthday_kids, today=today)


@app.route("/owner/classes", methods=["GET", "POST"])
@login_required
def owner_classes():
    if not require_owner():
        return redirect(url_for("unified_login"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if name:
            db.session.add(SchoolClass(nursery_id=current_user.nursery_id, name=name))
            db.session.commit()
        return redirect(url_for("owner_classes"))

    classes = SchoolClass.query.all()
    counts = {c.id: Child.query.filter_by(class_id=c.id).count() for c in classes}
    return render_template("owner_classes.html", classes=classes, counts=counts)


@app.route("/owner/classes/<int:class_id>/students", methods=["GET", "POST"])
@login_required
def owner_class_students(class_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    school_class = SchoolClass.query.get_or_404(class_id)

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        birth_date = request.form.get("birth_date") or None
        subscription_type = request.form.get("subscription_type", "full_time")
        monthly_fee = request.form.get("monthly_fee") or 0
        if name:
            db.session.add(Child(
                class_id=class_id, name=name, birth_date=birth_date,
                subscription_type=subscription_type, monthly_fee=monthly_fee,
            ))
            db.session.commit()
        return redirect(url_for("owner_class_students", class_id=class_id))

    show_archived = request.args.get("archived") == "1"
    students = Child.query.filter_by(class_id=class_id, archived=show_archived).all()
    parents_by_child = {}
    addons_by_child = {}
    total_due_by_child = {}
    for s in students:
        links = ParentChild.query.filter_by(child_id=s.id).all()
        parents_by_child[s.id] = [Parent.query.get(l.parent_id) for l in links]
        child_addons = ChildAddon.query.filter_by(child_id=s.id).all()
        addons_by_child[s.id] = [ca.addon for ca in child_addons]
        total_due_by_child[s.id] = compute_child_expected_fee(s)

    return render_template(
        "owner_students.html", school_class=school_class, students=students,
        parents_by_child=parents_by_child, addons_by_child=addons_by_child,
        total_due_by_child=total_due_by_child, today=date.today(), show_archived=show_archived,
    )


@app.route("/owner/child/<int:child_id>/archive", methods=["POST"])
@login_required
def owner_toggle_archive(child_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    child = Child.query.get_or_404(child_id)
    child.archived = not child.archived
    db.session.commit()
    return redirect(url_for("owner_class_students", class_id=child.class_id, archived="1" if child.archived else None))


@app.route("/owner/activities", methods=["GET", "POST"])
@login_required
def owner_activities():
    if not require_owner():
        return redirect(url_for("unified_login"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        cost = request.form.get("cost") or 0
        activity_date = request.form.get("date") or None
        notes = request.form.get("notes", "").strip()
        if name:
            db.session.add(Activity(
                nursery_id=current_user.nursery_id, name=name, cost=cost,
                date=activity_date or None, notes=notes,
            ))
            db.session.commit()
        return redirect(url_for("owner_activities"))

    activities = Activity.query.order_by(Activity.date.desc()).all()
    total_cost = sum(float(a.cost) for a in activities)
    return render_template("owner_activities.html", activities=activities, total_cost=total_cost)


@app.route("/owner/staff", methods=["GET", "POST"])
@login_required
def owner_staff():
    if not require_owner():
        return redirect(url_for("unified_login"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        role = request.form.get("role", "").strip()
        phone = request.form.get("phone", "").strip()
        salary = request.form.get("monthly_salary") or 0
        if name:
            db.session.add(Employee(
                nursery_id=current_user.nursery_id, name=name, role=role,
                phone=phone, monthly_salary=salary, hire_date=date.today(),
            ))
            db.session.commit()
        return redirect(url_for("owner_staff"))

    employees = Employee.query.filter_by(active=True).all()
    today = date.today()
    paid_status = {
        sp.employee_id: sp.paid for sp in
        SalaryPayment.query.filter_by(month=today.month, year=today.year).all()
    }
    methods = PaymentMethod.query.filter_by(active=True).all()

    advances_by_employee = {}
    for e in employees:
        outstanding = Advance.query.filter_by(employee_id=e.id, deducted=False).all()
        advances_by_employee[e.id] = {
            "items": outstanding,
            "total": sum(float(a.amount) for a in outstanding),
        }

    return render_template(
        "owner_staff.html", employees=employees, paid_status=paid_status,
        today=today, methods=methods, advances_by_employee=advances_by_employee,
    )


@app.route("/owner/staff/<int:employee_id>/pay-salary", methods=["POST"])
@login_required
def owner_pay_salary(employee_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    employee = Employee.query.get_or_404(employee_id)
    today = date.today()
    method_id = request.form.get("payment_method_id") or None

    outstanding_advances = Advance.query.filter_by(employee_id=employee_id, deducted=False).all()
    advances_total = sum(float(a.amount) for a in outstanding_advances)
    net_amount = max(float(employee.monthly_salary) - advances_total, 0)

    payment = SalaryPayment.query.filter_by(employee_id=employee_id, month=today.month, year=today.year).first()
    if not payment:
        payment = SalaryPayment(employee_id=employee_id, month=today.month, year=today.year, amount=net_amount)
        db.session.add(payment)
    payment.amount = net_amount
    payment.paid = True
    payment.paid_date = today
    payment.payment_method_id = method_id

    for adv in outstanding_advances:
        adv.deducted = True

    db.session.commit()
    return redirect(url_for("owner_staff"))


@app.route("/owner/staff/<int:employee_id>/advance", methods=["POST"])
@login_required
def owner_add_advance(employee_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    amount = request.form.get("amount")
    note = request.form.get("note", "").strip()
    if amount and float(amount) > 0:
        db.session.add(Advance(employee_id=employee_id, amount=amount, note=note))
        db.session.commit()
    return redirect(url_for("owner_staff"))


@app.route("/owner/parents", methods=["GET", "POST"])
@login_required
def owner_parents():
    if not require_owner():
        return redirect(url_for("unified_login"))

    if request.method == "POST":
        form_type = request.form.get("form_type")

        if form_type == "new_parent":
            name = request.form.get("name", "").strip()
            phone = request.form.get("phone", "").strip()
            password = request.form.get("password", "").strip()
            child_id = request.form.get("child_id")
            if name and phone and password:
                existing = Parent.query.filter_by(phone=phone).first()
                if existing:
                    return render_template_owner_parents_error("رقم الهاتف ده مسجل قبل كده لولي أمر تاني")
                new_parent = Parent(name=name, phone=phone)
                new_parent.set_password(password)
                db.session.add(new_parent)
                db.session.flush()
                if child_id:
                    db.session.add(ParentChild(parent_id=new_parent.id, child_id=int(child_id)))
                db.session.commit()

        elif form_type == "link_existing":
            parent_id = request.form.get("parent_id")
            child_id = request.form.get("child_id")
            if parent_id and child_id:
                exists = ParentChild.query.filter_by(parent_id=parent_id, child_id=child_id).first()
                if not exists:
                    db.session.add(ParentChild(parent_id=parent_id, child_id=child_id))
                    db.session.commit()

        return redirect(url_for("owner_parents"))

    parents = Parent.query.all()
    children = Child.query.all()
    links_by_parent = {}
    for p in parents:
        links = ParentChild.query.filter_by(parent_id=p.id).all()
        links_by_parent[p.id] = [Child.query.get(l.child_id) for l in links]

    return render_template("owner_parents.html", parents=parents, children=children, links_by_parent=links_by_parent)


def render_template_owner_parents_error(message):
    parents = Parent.query.all()
    children = Child.query.all()
    links_by_parent = {}
    for p in parents:
        links = ParentChild.query.filter_by(parent_id=p.id).all()
        links_by_parent[p.id] = [Child.query.get(l.child_id) for l in links]
    return render_template("owner_parents.html", parents=parents, children=children, links_by_parent=links_by_parent, error=message)


@app.route("/owner/addons", methods=["GET", "POST"])
@login_required
def owner_addons():
    if not require_owner():
        return redirect(url_for("unified_login"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        monthly_fee = request.form.get("monthly_fee") or 0
        if name:
            db.session.add(Addon(nursery_id=current_user.nursery_id, name=name, monthly_fee=monthly_fee))
            db.session.commit()
        return redirect(url_for("owner_addons"))

    addons = Addon.query.filter_by(active=True).all()
    subscriber_counts = {a.id: ChildAddon.query.filter_by(addon_id=a.id).count() for a in addons}
    return render_template("owner_addons.html", addons=addons, subscriber_counts=subscriber_counts)


@app.route("/owner/child/<int:child_id>/addons", methods=["GET", "POST"])
@login_required
def owner_child_addons(child_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    child = Child.query.get_or_404(child_id)

    if request.method == "POST":
        selected_ids = set(int(x) for x in request.form.getlist("addon_ids"))
        current_links = ChildAddon.query.filter_by(child_id=child_id).all()
        current_ids = {link.addon_id for link in current_links}

        for link in current_links:
            if link.addon_id not in selected_ids:
                db.session.delete(link)
        for addon_id in selected_ids - current_ids:
            db.session.add(ChildAddon(child_id=child_id, addon_id=addon_id))

        db.session.commit()
        return redirect(url_for("owner_class_students", class_id=child.class_id))

    all_addons = Addon.query.filter_by(active=True).all()
    active_ids = {ca.addon_id for ca in ChildAddon.query.filter_by(child_id=child_id).all()}
    return render_template("owner_child_addons.html", child=child, all_addons=all_addons, active_ids=active_ids)


@app.route("/owner/child/<int:child_id>/profile", methods=["GET", "POST"])
@login_required
def owner_child_profile(child_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    child = Child.query.get_or_404(child_id)

    if request.method == "POST":
        child.medical_notes = request.form.get("medical_notes", "").strip()
        child.discount_type = request.form.get("discount_type", "none")
        child.discount_value = request.form.get("discount_value") or 0

        photo_file = request.files.get("photo")
        if photo_file and photo_file.filename:
            raw = photo_file.read()
            if len(raw) > 3 * 1024 * 1024:
                return render_template("owner_child_profile.html", child=child, error="حجم الصورة كبير أوي، اختاري صورة أصغر من 3 ميجا")
            child.photo_data = base64.b64encode(raw).decode("ascii")
            child.photo_mime = photo_file.mimetype or "image/jpeg"

        db.session.commit()
        return redirect(url_for("owner_class_students", class_id=child.class_id))

    return render_template("owner_child_profile.html", child=child, error=None)


@app.route("/child/<int:child_id>/photo")
@login_required
def child_photo(child_id):
    child = Child.query.get_or_404(child_id)

    if current_role() == "parent" and not child_belongs_to_parent(child_id, current_user.id):
        return "غير مصرح", 403

    if not child.photo_data:
        return "", 404

    raw = base64.b64decode(child.photo_data)
    return app.response_class(raw, mimetype=child.photo_mime or "image/jpeg")


@app.route("/owner/payment-methods", methods=["GET", "POST"])
@login_required
def owner_payment_methods():
    if not require_owner():
        return redirect(url_for("unified_login"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if name:
            db.session.add(PaymentMethod(nursery_id=current_user.nursery_id, name=name))
            db.session.commit()
        return redirect(url_for("owner_payment_methods"))

    methods = PaymentMethod.query.filter_by(active=True).all()
    return render_template("owner_payment_methods.html", methods=methods)


@app.route("/owner/payment-methods/<int:method_id>/edit", methods=["POST"])
@login_required
def owner_edit_payment_method(method_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    method = PaymentMethod.query.get_or_404(method_id)
    name = request.form.get("name", "").strip()
    if name:
        method.name = name
        db.session.commit()
    return redirect(url_for("owner_payment_methods"))


@app.route("/owner/payment-methods/<int:method_id>/delete", methods=["POST"])
@login_required
def owner_delete_payment_method(method_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    method = PaymentMethod.query.get_or_404(method_id)
    method.active = False
    db.session.commit()
    return redirect(url_for("owner_payment_methods"))


def compute_payment_statement(start, end):
    methods = PaymentMethod.query.filter_by(active=True).all()
    summary = {m.id: {"name": m.name, "in": 0.0, "out": 0.0} for m in methods}
    summary["none"] = {"name": "بدون طريقة محددة", "in": 0.0, "out": 0.0}

    transactions = []

    tuition_payments = Payment.query.filter(
        Payment.paid.is_(True), Payment.paid_date >= start, Payment.paid_date <= end
    ).all()
    for p in tuition_payments:
        key = p.payment_method_id or "none"
        summary.setdefault(key, {"name": p.payment_method.name if p.payment_method else "بدون طريقة محددة", "in": 0.0, "out": 0.0})
        summary[key]["in"] += float(p.amount)
        transactions.append({
            "date": p.paid_date, "type": "تحصيل اشتراك", "detail": p.child.name,
            "amount": float(p.amount), "direction": "in",
            "method": p.payment_method.name if p.payment_method else "—",
        })

    expenses = Expense.query.filter(Expense.date >= start, Expense.date <= end).all()
    for e in expenses:
        key = e.payment_method_id or "none"
        summary.setdefault(key, {"name": e.payment_method.name if e.payment_method else "بدون طريقة محددة", "in": 0.0, "out": 0.0})
        summary[key]["out"] += float(e.amount)
        transactions.append({
            "date": e.date, "type": f"مصروف — {e.category}", "detail": e.description or "",
            "amount": float(e.amount), "direction": "out",
            "method": e.payment_method.name if e.payment_method else "—",
        })

    salary_payments = SalaryPayment.query.filter(
        SalaryPayment.paid.is_(True), SalaryPayment.paid_date >= start, SalaryPayment.paid_date <= end
    ).all()
    for sp in salary_payments:
        key = sp.payment_method_id or "none"
        summary.setdefault(key, {"name": sp.payment_method.name if sp.payment_method else "بدون طريقة محددة", "in": 0.0, "out": 0.0})
        summary[key]["out"] += float(sp.amount)
        transactions.append({
            "date": sp.paid_date, "type": "مرتب", "detail": sp.employee.name,
            "amount": float(sp.amount), "direction": "out",
            "method": sp.payment_method.name if sp.payment_method else "—",
        })

    transactions.sort(key=lambda t: t["date"], reverse=True)
    summary_rows = [v for k, v in summary.items() if v["in"] > 0 or v["out"] > 0]
    return summary_rows, transactions


@app.route("/owner/payment-methods/statement/export")
@login_required
def owner_payment_statement_export():
    if not require_owner():
        return redirect(url_for("unified_login"))

    today = date.today()
    start = request.args.get("start") or today.replace(day=1).isoformat()
    end = request.args.get("end") or today.isoformat()
    summary_rows, transactions = compute_payment_statement(start, end)

    wb = Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True
    ws.title = "ملخص الطرق"

    ws.append([f"كشف حركة طرق التحصيل — من {start} إلى {end}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["الطريقة", "تحصيل (داخل)", "صرف (خارج)", "الصافي"])
    style_header_row(ws, ws.max_row, 4)
    for row in summary_rows:
        ws.append([row["name"], row["in"], row["out"], row["in"] - row["out"]])
    for col, width in zip("ABCD", [22, 16, 16, 16]):
        ws.column_dimensions[col].width = width

    ws2 = wb.create_sheet("كل الحركات")
    ws2.sheet_view.rightToLeft = True
    ws2.append(["التاريخ", "النوع", "التفاصيل", "الطريقة", "المبلغ"])
    style_header_row(ws2, 1, 5)
    for t in transactions:
        signed_amount = t["amount"] if t["direction"] == "in" else -t["amount"]
        ws2.append([t["date"].strftime("%d/%m/%Y") if t["date"] else "", t["type"], t["detail"], t["method"], signed_amount])
    for col, width in zip("ABCDE", [14, 22, 26, 18, 14]):
        ws2.column_dimensions[col].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return app.response_class(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=كشف-حركة-{start}-الى-{end}.xlsx"},
    )


@app.route("/owner/payment-methods/statement")
@login_required
def owner_payment_methods_statement():
    if not require_owner():
        return redirect(url_for("unified_login"))

    today = date.today()
    start = request.args.get("start") or today.replace(day=1).isoformat()
    end = request.args.get("end") or today.isoformat()

    summary_rows, transactions = compute_payment_statement(start, end)

    return render_template(
        "owner_payment_statement.html", start=start, end=end,
        summary_rows=summary_rows, transactions=transactions,
    )


@app.route("/owner/attendance-report")
@login_required
def owner_attendance_report():
    if not require_owner():
        return redirect(url_for("unified_login"))

    month = int(request.args.get("month", date.today().month))
    year = int(request.args.get("year", date.today().year))

    records = AttendanceRecord.query.filter(
        db.extract("month", AttendanceRecord.date) == month,
        db.extract("year", AttendanceRecord.date) == year,
    ).all()

    school_days = len({r.date for r in records}) or 1

    by_child = {}
    for r in records:
        by_child.setdefault(r.child_id, {"present": 0, "child": r.child})
        if r.present:
            by_child[r.child_id]["present"] += 1

    rows = []
    for child_id, data in by_child.items():
        pct = round((data["present"] / school_days) * 100)
        rows.append({"child": data["child"], "present": data["present"], "school_days": school_days, "pct": pct, "low": pct < 80})
    rows.sort(key=lambda r: r["pct"])

    return render_template("owner_attendance_report.html", rows=rows, month=month, year=year, school_days=school_days)


@app.route("/owner/dues")
@login_required
def owner_dues():
    if not require_owner():
        return redirect(url_for("unified_login"))

    unpaid = Payment.query.filter_by(paid=False).order_by(Payment.year, Payment.month).all()

    by_child = {}
    for p in unpaid:
        if p.child.archived:
            continue
        by_child.setdefault(p.child_id, {"child": p.child, "items": [], "total": 0.0})
        by_child[p.child_id]["items"].append(p)
        by_child[p.child_id]["total"] += float(p.amount)

    rows = sorted(by_child.values(), key=lambda r: r["total"], reverse=True)
    grand_total = sum(r["total"] for r in rows)

    return render_template("owner_dues.html", rows=rows, grand_total=grand_total)


@app.route("/owner/tuition", methods=["GET", "POST"])
@login_required
def owner_tuition():
    if not require_owner():
        return redirect(url_for("unified_login"))

    month = int(request.args.get("month", date.today().month))
    year = int(request.args.get("year", date.today().year))

    if request.method == "POST":
        child_id = int(request.form.get("child_id"))
        amount = request.form.get("amount")
        method_id = request.form.get("payment_method_id") or None

        payment = Payment.query.filter_by(child_id=child_id, month=month, year=year).first()
        if not payment:
            payment = Payment(child_id=child_id, month=month, year=year, amount=amount)
            db.session.add(payment)
        payment.amount = amount
        payment.paid = True
        payment.paid_date = date.today()
        payment.payment_method_id = method_id
        db.session.commit()
        return redirect(url_for("owner_tuition", month=month, year=year))

    children = Child.query.filter_by(archived=False).all()
    existing_payments = {
        p.child_id: p for p in Payment.query.filter_by(month=month, year=year).all()
    }
    methods = PaymentMethod.query.filter_by(active=True).all()

    rows = []
    for c in children:
        expected = compute_child_expected_fee(c)
        rows.append({
            "child": c,
            "expected": expected,
            "payment": existing_payments.get(c.id),
        })

    return render_template("owner_tuition.html", rows=rows, methods=methods, month=month, year=year)


@app.route("/owner/expense-categories", methods=["GET", "POST"])
@login_required
def owner_expense_categories():
    if not require_owner():
        return redirect(url_for("unified_login"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if name:
            db.session.add(ExpenseCategory(nursery_id=current_user.nursery_id, name=name))
            db.session.commit()
        return redirect(url_for("owner_expense_categories"))

    categories = ExpenseCategory.query.filter_by(active=True).all()
    return render_template("owner_expense_categories.html", categories=categories)


@app.route("/owner/expense-categories/<int:category_id>/edit", methods=["POST"])
@login_required
def owner_edit_expense_category(category_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    category = ExpenseCategory.query.get_or_404(category_id)
    name = request.form.get("name", "").strip()
    if name:
        category.name = name
        db.session.commit()
    return redirect(url_for("owner_expense_categories"))


@app.route("/owner/expense-categories/<int:category_id>/delete", methods=["POST"])
@login_required
def owner_delete_expense_category(category_id):
    if not require_owner():
        return redirect(url_for("unified_login"))

    category = ExpenseCategory.query.get_or_404(category_id)
    category.active = False
    db.session.commit()
    return redirect(url_for("owner_expense_categories"))


@app.route("/owner/expenses", methods=["GET", "POST"])
@login_required
def owner_expenses():
    if not require_owner():
        return redirect(url_for("unified_login"))

    if request.method == "POST":
        category = request.form.get("category", "").strip()
        description = request.form.get("description", "").strip()
        amount = request.form.get("amount") or 0
        expense_date = request.form.get("date") or date.today().isoformat()
        method_id = request.form.get("payment_method_id") or None
        if category and float(amount) > 0:
            db.session.add(Expense(
                nursery_id=current_user.nursery_id, category=category, description=description,
                amount=amount, date=expense_date, created_by_owner_id=current_user.id,
                payment_method_id=method_id,
            ))
            db.session.commit()
        return redirect(url_for("owner_expenses"))

    expenses = Expense.query.order_by(Expense.date.desc()).all()
    total = sum(float(e.amount) for e in expenses)
    methods = PaymentMethod.query.filter_by(active=True).all()
    categories = ExpenseCategory.query.filter_by(active=True).all()
    return render_template("owner_expenses.html", expenses=expenses, total=total, methods=methods, categories=categories)


@app.route("/owner/reports")
@login_required
def owner_reports():
    if not require_owner():
        return redirect(url_for("unified_login"))

    month = int(request.args.get("month", date.today().month))
    year = int(request.args.get("year", date.today().year))

    tuition_paid = db.session.query(db.func.coalesce(db.func.sum(Payment.amount), 0)).filter(
        Payment.month == month, Payment.year == year, Payment.paid.is_(True)
    ).scalar()
    tuition_due = db.session.query(db.func.coalesce(db.func.sum(Payment.amount), 0)).filter(
        Payment.month == month, Payment.year == year, Payment.paid.is_(False)
    ).scalar()

    expenses = Expense.query.filter(
        db.extract("month", Expense.date) == month, db.extract("year", Expense.date) == year
    ).all()
    expenses_by_category = {}
    for e in expenses:
        expenses_by_category[e.category] = expenses_by_category.get(e.category, 0) + float(e.amount)
    expenses_total = sum(expenses_by_category.values())

    salaries_total = db.session.query(db.func.coalesce(db.func.sum(SalaryPayment.amount), 0)).filter(
        SalaryPayment.month == month, SalaryPayment.year == year, SalaryPayment.paid.is_(True)
    ).scalar()

    activities = Activity.query.filter(
        db.extract("month", Activity.date) == month, db.extract("year", Activity.date) == year
    ).all()
    activities_cost = sum(float(a.cost) for a in activities)

    net = float(tuition_paid) - expenses_total - float(salaries_total) - activities_cost

    return render_template(
        "owner_reports.html", month=month, year=year,
        tuition_paid=tuition_paid, tuition_due=tuition_due,
        expenses_by_category=expenses_by_category, expenses_total=expenses_total,
        salaries_total=salaries_total, activities_cost=activities_cost, net=net,
    )


def style_header_row(ws, row_num, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5D50", end_color="2F5D50", fill_type="solid")
        cell.alignment = Alignment(horizontal="right")


@app.route("/owner/backup/export")
@login_required
def owner_backup_export():
    if not require_owner():
        return redirect(url_for("unified_login"))

    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        ws.sheet_view.rightToLeft = True
        ws.append(headers)
        style_header_row(ws, 1, len(headers))
        for row in rows:
            ws.append(row)
        for i, h in enumerate(headers):
            ws.column_dimensions[chr(65 + i)].width = max(14, len(str(h)) + 4)

    children = Child.query.all()
    add_sheet("الأطفال", ["الاسم", "الفصل", "تاريخ الميلاد", "نوع الاشتراك", "الاشتراك الشهري", "مؤرشف"], [
        [c.name, c.school_class.name, c.birth_date.isoformat() if c.birth_date else "", c.subscription_type, float(c.monthly_fee), "نعم" if c.archived else "لا"]
        for c in children
    ])

    payments = Payment.query.order_by(Payment.year, Payment.month).all()
    add_sheet("المصروفات الدراسية", ["الطفل", "الشهر", "السنة", "المبلغ", "مدفوع", "تاريخ السداد", "الطريقة"], [
        [p.child.name, p.month, p.year, float(p.amount), "نعم" if p.paid else "لا",
         p.paid_date.isoformat() if p.paid_date else "", p.payment_method.name if p.payment_method else ""]
        for p in payments
    ])

    expenses = Expense.query.order_by(Expense.date).all()
    add_sheet("المصروفات العامة", ["النوع", "الوصف", "المبلغ", "التاريخ", "الطريقة"], [
        [e.category, e.description or "", float(e.amount), e.date.isoformat(), e.payment_method.name if e.payment_method else ""]
        for e in expenses
    ])

    employees = Employee.query.all()
    add_sheet("الموظفين", ["الاسم", "الوظيفة", "الهاتف", "المرتب الشهري", "نشط"], [
        [e.name, e.role or "", e.phone or "", float(e.monthly_salary), "نعم" if e.active else "لا"]
        for e in employees
    ])

    salaries = SalaryPayment.query.order_by(SalaryPayment.year, SalaryPayment.month).all()
    add_sheet("المرتبات المدفوعة", ["الموظف", "الشهر", "السنة", "المبلغ", "مدفوع", "الطريقة"], [
        [s.employee.name, s.month, s.year, float(s.amount), "نعم" if s.paid else "لا", s.payment_method.name if s.payment_method else ""]
        for s in salaries
    ])

    trips = Trip.query.all()
    add_sheet("الرحلات", ["الاسم", "التاريخ", "السعر النهائي", "الحالة", "عدد المشتركين"], [
        [t.title, t.date.isoformat() if t.date else "", float(t.final_price or 0), t.status, len(t.registrations)]
        for t in trips
    ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return app.response_class(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=نسخة-احتياطية-{date.today().isoformat()}.xlsx"},
    )


@app.route("/owner/reports/export")
@login_required
def owner_reports_export():
    if not require_owner():
        return redirect(url_for("unified_login"))

    month = int(request.args.get("month", date.today().month))
    year = int(request.args.get("year", date.today().year))

    tuition_paid = db.session.query(db.func.coalesce(db.func.sum(Payment.amount), 0)).filter(
        Payment.month == month, Payment.year == year, Payment.paid.is_(True)
    ).scalar()
    tuition_due = db.session.query(db.func.coalesce(db.func.sum(Payment.amount), 0)).filter(
        Payment.month == month, Payment.year == year, Payment.paid.is_(False)
    ).scalar()
    expenses = Expense.query.filter(
        db.extract("month", Expense.date) == month, db.extract("year", Expense.date) == year
    ).all()
    expenses_by_category = {}
    for e in expenses:
        expenses_by_category[e.category] = expenses_by_category.get(e.category, 0) + float(e.amount)
    salaries_total = float(db.session.query(db.func.coalesce(db.func.sum(SalaryPayment.amount), 0)).filter(
        SalaryPayment.month == month, SalaryPayment.year == year, SalaryPayment.paid.is_(True)
    ).scalar())
    activities = Activity.query.filter(
        db.extract("month", Activity.date) == month, db.extract("year", Activity.date) == year
    ).all()
    activities_cost = sum(float(a.cost) for a in activities)
    expenses_total = sum(expenses_by_category.values())
    net = float(tuition_paid) - expenses_total - salaries_total - activities_cost

    wb = Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True
    ws.title = f"تقرير {month}-{year}"

    ws.append([f"التقرير المالي — {month}/{year}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["البند", "المبلغ (جنيه)"])
    style_header_row(ws, ws.max_row, 2)
    ws.append(["مصروفات دراسية محصّلة", float(tuition_paid)])
    ws.append(["مصروفات دراسية مستحقة", float(tuition_due)])
    ws.append([])
    ws.append(["المصروفات حسب النوع", ""])
    style_header_row(ws, ws.max_row, 2)
    for cat, amount in expenses_by_category.items():
        ws.append([cat, amount])
    ws.append(["إجمالي المصروفات العامة", expenses_total])
    ws.append(["مرتبات مدفوعة", salaries_total])
    ws.append(["تكلفة الأنشطة", activities_cost])
    ws.append([])
    ws.append(["صافي الربح", net])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return app.response_class(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=تقرير-{month}-{year}.xlsx"},
    )


# ---------------------------------------------------------------------------
# CLI helper: seed demo data
# ---------------------------------------------------------------------------

@app.cli.command("db-init")
def db_init():
    """Creates tables without wiping existing data. Safe to run on every deploy."""
    with app.app_context():
        db.create_all()
    print("تم التأكد من وجود الجداول.")


@app.route("/internal/send-due-reminders")
def http_send_due_reminders():
    """Triggered by a free external scheduler (e.g. cron-job.org) — sends a push
    notification to parents of children with unpaid tuition from a past month."""
    key = request.args.get("key")
    expected_key = os.environ.get("INTERNAL_TASK_KEY")
    if not expected_key or key != expected_key:
        return "غير مصرح", 403

    today = date.today()
    overdue = Payment.query.filter_by(paid=False).all()
    sent = 0
    for p in overdue:
        # Only remind for months strictly before the current one (truly overdue,
        # not this month's not-yet-collected payment).
        if (p.year, p.month) >= (today.year, today.month):
            continue
        if p.child.archived:
            continue
        notify_child_parents(
            p.child_id,
            f"تذكير بمصروفات {p.child.name}",
            f"مصروفات شهر {p.month}/{p.year} ({p.amount} ج) لسه متسددتش، برجاء التواصل مع الحضانة",
        )
        sent += 1

    return {"reminders_sent": sent}


@app.route("/internal/generate-monthly-tuition")
def http_generate_monthly_tuition():
    """Triggered by a free external scheduler (e.g. cron-job.org) on day 1 of each
    month. Protected by a secret key so it can't be called by anyone else."""
    key = request.args.get("key")
    expected_key = os.environ.get("INTERNAL_TASK_KEY")
    if not expected_key or key != expected_key:
        return "غير مصرح", 403

    today = date.today()
    children = Child.query.filter_by(archived=False).all()
    created = 0
    for child in children:
        existing = Payment.query.filter_by(child_id=child.id, month=today.month, year=today.year).first()
        if existing:
            continue
        expected = compute_child_expected_fee(child)
        db.session.add(Payment(child_id=child.id, month=today.month, year=today.year, amount=expected, paid=False))
        created += 1
    db.session.commit()
    return {"created": created, "month": today.month, "year": today.year}


@app.cli.command("generate-monthly-tuition")
def generate_monthly_tuition():
    """Creates this month's tuition Payment record for every active (non-archived) child,
    based on their current subscription fee + addons - discount. Meant to run on day 1
    of each month via a scheduled cron job. Safe to re-run: skips children who already
    have a payment record for the month."""
    with app.app_context():
        today = date.today()
        children = Child.query.filter_by(archived=False).all()
        created = 0
        for child in children:
            existing = Payment.query.filter_by(child_id=child.id, month=today.month, year=today.year).first()
            if existing:
                continue
            expected = compute_child_expected_fee(child)
            db.session.add(Payment(
                child_id=child.id, month=today.month, year=today.year,
                amount=expected, paid=False,
            ))
            created += 1
        db.session.commit()
        print(f"تم إنشاء {created} سجل مصاريف لشهر {today.month}/{today.year}")


@app.cli.command("seed")
def seed():
    """Run with: flask --app app.py seed"""
    db.drop_all()
    db.create_all()

    nursery = Nursery(name="حضانة البراعم الصغيرة")
    db.session.add(nursery)
    db.session.flush()

    school_class = SchoolClass(nursery_id=nursery.id, name="الفراشات")
    db.session.add(school_class)
    db.session.flush()

    teacher = Teacher(nursery_id=nursery.id, name="مروة أحمد", phone="01000000001", class_id=school_class.id)
    teacher.set_password("teacher123")
    db.session.add(teacher)

    parent = Parent(name="ولية أمر يوسف", phone="01000000002")
    parent.set_password("parent123")
    db.session.add(parent)
    db.session.flush()

    child = Child(class_id=school_class.id, name="يوسف أحمد")
    db.session.add(child)
    db.session.flush()

    db.session.add(ParentChild(parent_id=parent.id, child_id=child.id))

    owner = Owner(nursery_id=nursery.id, name="مالكة الحضانة", phone="01000000009")
    owner.set_password("owner123")
    db.session.add(owner)

    db.session.add(DailyLog(
        child_id=child.id, date=date.today(),
        meal_status="اتاكلت كلها", nap_minutes=120, mood="😊",
        note="رسم لوحة جميلة وقال إنها هدية لماما",
        created_by_teacher_id=teacher.id,
    ))
    db.session.add(AttendanceRecord(child_id=child.id, date=date.today(), present=True))
    db.session.add(Payment(child_id=child.id, month=date.today().month, year=date.today().year, amount=2500, paid=True))

    db.session.commit()
    print("تم إنشاء بيانات تجريبية:")
    print("  معلمة: 01000000001 / teacher123")
    print("  ولية أمر: 01000000002 / parent123")


if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    app.run(debug=True, host="0.0.0.0", port=5000)
